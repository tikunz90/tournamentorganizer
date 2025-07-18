from django.db.models.query import Prefetch
from openpyxl.cell.cell import TYPE_STRING
from beachhandball_app import helper
from beachhandball_app.models.Tournaments import Tournament, TournamentSettings
from beachhandball_app.models.choices import CATEGORY_CHOICES
from beachhandball_app.models.Game import Game
from beachhandball_app.models.Player import Player, PlayerStats

import os
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from django.conf import settings

def create_pregame_report_excel(game):
    print('ENTER create_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)

    game = (
        Game.objects
        .select_related(
            "tournament",
            "tournament_event__category",
            "tournament_state__tournament_stage",
            "court",
            "team_st_a__team",
            "team_st_b__team",
            "ref_a",
            "ref_b"
        )
        .prefetch_related(
            Prefetch("team_st_a__team__player_set", queryset=Player.objects.order_by('number'), to_attr="players_a"),
            Prefetch("team_st_b__team__player_set", queryset=Player.objects.order_by('number'), to_attr="players_b"),
            Prefetch("team_st_a__team__coach_set", to_attr="coaches_a"),
            Prefetch("team_st_b__team__coach_set", to_attr="coaches_b"),
        )
        .get(pk=game.pk)
    )

    tsettings = TournamentSettings.objects.get(tournament=game.tournament_shared)
    if not tsettings.game_report_template:
        return ('', '')

    gr_template = tsettings.game_report_template
    filename_template = gr_template.filename #
    filename = 'PreGame_' + game.tournament_event.category.abbreviation + str(game.id_counter) + '.xlsx'

    fullfilepath_template = os.path.join(settings.GAME_REPORT_DIR, filename_template)
    fullfilepath_report = os.path.join(settings.GAME_REPORT_DIR, filename)
    # copy template
    copyfile(fullfilepath_template, fullfilepath_report)

    if os.path.isfile(fullfilepath_report):
        print('file exists')
        wb = load_workbook(filename = fullfilepath_report)
        ws = wb.active
        #img = Image(os.path.join(settings.GAME_REPORT_DIR,'dhb_logo.png'))
        #ws_copy = wb.copy_worksheet(ws)
        #ws.add_image(img, 'T1')
        # game id
        tsettings = TournamentSettings.objects.get(tournament=game.tournament_shared)
        tstage = game.tournament_state.tournament_stage
        ws[gr_template.cell_game_id] = game.tournament_event.category.abbreviation + str(game.id_counter)
        
        # category
        if game.tournament_event.category.abbreviation == 'M': # CATEGORY_CHOICES[0][0]:
            ws[gr_template.cell_category_men] = 'X'
        elif game.tournament_event.category.abbreviation == 'W': # CATEGORY_CHOICES[1][0]:
            ws[gr_template.cell_category_women] = 'X'

        # court
        ws[gr_template.cell_court_number] = game.court.number

        # team names
        ws[gr_template.cell_team_a_name] = game.team_st_a.team.name
        ws[gr_template.cell_team_b_name] = game.team_st_b.team.name

        # ref names
        if game.ref_a is not None:
            ws[gr_template.cell_ref_a_name] = game.ref_a.name + ', ' + game.ref_a.first_name
        if game.ref_b is not None:
            ws[gr_template.cell_ref_b_name] = game.ref_b.name + ', ' + game.ref_b.first_name 

        # stage
        if tstage is not None:
            if tstage.tournament_stage == 'GROUP_STAGE':
                ws[gr_template.cell_group_stage] = 'X'
            if tstage.tournament_stage == 'KNOCKOUT_STAGE' or tstage.tournament_stage == 'FINAL' or tstage.tournament_stage == 'PLAYOFF_STAGE':
                ws[gr_template.cell_knockout_stage] = 'X'

        # Date and Time
        date = game.starttime.strftime("%d.%m.%Y")
        time = game.starttime.strftime("%H:%M")
        ws[gr_template.cell_date] = date
        ws[gr_template.cell_time] = time

        # Player
        max_num_player = tsettings.amount_players_report
        row_start_a = gr_template.team_a_start_row
        act_player_counter = 1
        for player in game.team_st_a.team.player_set.order_by('number').all():
            if act_player_counter > max_num_player:
                break
            if not player.is_active:
                continue
            ws[gr_template.team_a_player_number_col+str(row_start_a)]=player.number
            ws[gr_template.team_a_player_name_col+str(row_start_a)]=player.name + ", " + player.first_name + " (" + str(player.id) + ")"
            act_player_counter = act_player_counter + 1
            row_start_a = row_start_a + 1
        row_start_b = gr_template.team_b_start_row
        act_player_counter = 1
        for player in game.team_st_b.team.player_set.order_by('number').all():
            if act_player_counter > max_num_player:
                break
            if not player.is_active:
                continue
            ws[gr_template.team_b_player_number_col+str(row_start_b)]=player.number
            ws[gr_template.team_b_player_name_col+str(row_start_b)]=player.name + ", " + player.first_name + " (" + str(player.id) + ")"
            act_player_counter = act_player_counter + 1
            row_start_b = row_start_b + 1

        # Coaches
        max_num_coaches = gr_template.max_num_coaches
        row_start_coach_a = gr_template.team_a_start_row_coaches
        for coach in game.team_st_a.team.coach_set.all()[:max_num_coaches]:
            ws[gr_template.team_a_coach_name_col +str(row_start_coach_a)]=coach.name + ", " + coach.first_name
            row_start_coach_a = row_start_coach_a + 1
        row_start_coach_b = gr_template.team_b_start_row_coaches
        for coach in game.team_st_b.team.coach_set.all()[:max_num_coaches]:
            ws[gr_template.team_b_coach_name_col+str(row_start_coach_b)]=coach.name + ", " + coach.first_name
            row_start_coach_b = row_start_coach_b + 1

        wb.save(fullfilepath_report)

        return fullfilepath_report, filename
    return ('', '')


def create_all_tstate_pregame_report_excel(tstate):
    print('ENTER create_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)

    tsettings = TournamentSettings.objects.get(tournament=tstate.tournament_event.tournament)
    if not tsettings.game_report_template:
        return ''
    
    gr_template = tsettings.game_report_template

    filename_template = tsettings.game_report_template.filename
    filename = 'PreGame_ALL_'  + str(tstate.name) + '.xlsx'

    fullfilepath_template = os.path.join(settings.GAME_REPORT_DIR, filename_template)
    fullfilepath_report = os.path.join(settings.GAME_REPORT_DIR, filename)
    # copy template
    copyfile(fullfilepath_template, fullfilepath_report)

    if os.path.isfile(fullfilepath_report):
        print('file exists')
        tsettings = TournamentSettings.objects.get(tournament=tstate.tournament_event.tournament)
        games = Game.objects.filter(tournament_state=tstate)
        tstage = tstate.tournament_stage
        wb = load_workbook(filename = fullfilepath_report)
        ws_origin = wb.active
        #img = Image(os.path.join(settings.GAME_REPORT_DIR,'dhb_logo.png'))
        for game in games:
            
            ws_game = wb.copy_worksheet(ws_origin)
            ws_game.title = game.tournament_event.category.abbreviation + str(game.id)
            # game id
            ws_game[gr_template.cell_game_id] = game.tournament_event.category.abbreviation + str(game.id)
            
            # category
            if game.tournament_event.category.abbreviation == 'M': # CATEGORY_CHOICES[0][0]:
                ws_game[gr_template.cell_category_men] = 'X'
            elif game.tournament_event.category.abbreviation == 'W': # CATEGORY_CHOICES[1][0]:
                ws_game[gr_template.cell_category_women] = 'X'

            # court
            ws_game[gr_template.cell_court_number] = game.court.number

            # team names
            ws_game[gr_template.cell_team_a_name] = game.team_st_a.team.name
            ws_game[gr_template.cell_team_b_name] = game.team_st_b.team.name

            # ref names
            if game.ref_a is not None:
                ws_game[gr_template.cell_ref_a_name] = game.ref_a.name + ', ' + game.ref_a.first_name
            if game.ref_b is not None:
                ws_game[gr_template.cell_ref_b_name] = game.ref_b.name + ', ' + game.ref_b.first_name 

            # stage
            if tstage is not None:
                if tstage.tournament_stage == 'GROUP_STAGE':
                    ws_game[gr_template.cell_group_stage] = 'X'
                if tstage.tournament_stage == 'KNOCKOUT_STAGE' or tstage.tournament_stage == 'FINAL' or tstage.tournament_stage == 'PLAYOFF_STAGE':
                    ws_game[gr_template.cell_knockout_stage] = 'X'

            # Date and Time
            date = game.starttime.strftime("%d.%m.%Y")
            time = game.starttime.strftime("%H:%M")
            ws_game[gr_template.cell_date] = date
            ws_game[gr_template.cell_time] = time
            max_num_player = tsettings.amount_players_report
            row_start_a = gr_template.team_a_start_row
            act_player_counter = 1
            for player in game.team_st_a.team.player_set.all():
                if act_player_counter > max_num_player:
                    break
                if not player.is_active:
                    continue
                ws_game[gr_template.team_a_player_number_col+str(row_start_a)]=player.number
                ws_game[gr_template.team_a_player_name_col  +str(row_start_a)]=player.name + ", " + player.first_name
                act_player_counter = act_player_counter + 1
                row_start_a = row_start_a + 1
            row_start_b = gr_template.team_b_start_row
            act_player_counter = 1
            for player in game.team_st_b.team.player_set.all():
                if act_player_counter > max_num_player:
                    break
                if not player.is_active:
                    continue
                ws_game[gr_template.team_b_player_number_col+str(row_start_b)]=player.number
                ws_game[gr_template.team_b_player_name_col  +str(row_start_b)]=player.name + ", " + player.first_name
                act_player_counter = act_player_counter + 1
                row_start_b = row_start_b + 1
            max_num_coaches = gr_template.max_num_coaches
            row_start_coach_a = gr_template.team_a_start_row_coaches
            for coach in game.team_st_a.team.coach_set.all()[:max_num_coaches]:
                ws_game[gr_template.team_a_coach_name_col+str(row_start_coach_a)]=coach.name + ", " + coach.first_name
                row_start_coach_a = row_start_coach_a + 1
            row_start_coach_b = gr_template.team_b_start_row_coaches
            for coach in game.team_st_b.team.coach_set.all()[:max_num_coaches]:
                ws_game[gr_template.team_b_coach_name_col+str(row_start_coach_b)]=coach.name + ", " + coach.first_name
                row_start_coach_b = row_start_coach_b + 1
        wb.remove_sheet(ws_origin)
        wb.save(fullfilepath_report)

        return fullfilepath_report, filename
    return ''

def import_game_report_excel(game):
    print('ENTER import_game_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)
    tourn = Tournament.objects.get(id=28)
    tsettings = TournamentSettings.objects.get(tournament=tourn)
    if not tsettings.game_report_template:
        return ''

    gr_template = tsettings.game_report_template


def pre_import_single_game_report(game, filename):
    print('ENTER import_single_game_report ' + filename)
    result = {'isError': False, 'msg': 'OK', 'playerstats_a': [], 'playerstats_b': [], 'score_team_a_halftime_1': 0, 'score_team_a_halftime_2': 0, 'score_team_a_penalty': 0, 'score_team_b_halftime_1': 0, 'score_team_b_halftime_2': 0, 'score_team_b_penalty': 0}

    tsettings = TournamentSettings.objects.get(tournament=game.tournament)
    if not tsettings.game_report_template:
        result['isError'] = True
        result['msg'] = 'No template defined'
        return result
    tmp = tsettings.game_report_template
    
    global_ps_a = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_a, is_ranked=True).all()]
    global_ps_b = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_b, is_ranked=True).all()]

    local_ps_a = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, game=game, player__team=game.team_a, is_ranked=False).all()]
    local_ps_b = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, game=game, player__team=game.team_b, is_ranked=False).all()]
    
    wb = load_workbook(filename = filename, data_only=True)
    ws = wb.active

    result['score_team_a_halftime_1'] = ws[tmp.cell_score_team_a_halftime_1].value
    result['score_team_a_halftime_2'] = ws[tmp.cell_score_team_a_halftime_2].value
    result['score_team_a_penalty'] = ws[tmp.cell_score_team_a_penalty].value
    result['score_team_b_halftime_1'] = ws[tmp.cell_score_team_b_halftime_1].value
    result['score_team_b_halftime_2'] = ws[tmp.cell_score_team_b_halftime_2].value
    result['score_team_b_penalty'] = ws[tmp.cell_score_team_b_penalty].value

    if result['score_team_a_halftime_1'] is None:
        result['score_team_a_halftime_1'] = 0
    if result['score_team_a_halftime_2'] is None:
        result['score_team_a_halftime_2'] = 0
    if result['score_team_a_penalty'] is None:
        result['score_team_a_penalty'] = 0
    if result['score_team_b_halftime_1'] is None:
        result['score_team_b_halftime_1'] = 0
    if result['score_team_b_halftime_2'] is None:
        result['score_team_b_halftime_2'] = 0
    if result['score_team_b_penalty'] is None:
        result['score_team_b_penalty'] = 0

    if int(result['score_team_a_halftime_1']) > 99 or int(result['score_team_a_halftime_2']) > 99 or int(result['score_team_a_penalty']) > 99:
        result['isError'] = True
        result['msg'] = 'Double check result. score is too high...'
        return result
    
    if int(result['score_team_b_halftime_1']) > 99 or int(result['score_team_b_halftime_2']) > 99 or int(result['score_team_b_penalty']) > 99:
        result['isError'] = True
        result['msg'] = 'Double check result. score is too high...'
        return result

    messages = []
    maxPoints = 50
    for iRow in range(0,tsettings.amount_players_report):
        actRow = iRow + tmp.team_a_start_row
        actNumber = ws[tmp.team_a_player_number_col+str(actRow)].value
        if actNumber is None:
            continue
        actName = ws[tmp.team_a_player_name_col+str(actRow)].value
        if actName is None or actName == '':
            continue
        actPoints = ws[tmp.team_a_player_points_col+str(actRow)].value
        if actPoints is None:
            actPoints = 0
        if actPoints > maxPoints:
            messages.append("#" + str(actNumber) + " score too high...")
            continue
        player_id, res = intTryParse(actName[actName.find("(")+1:actName.find(")")])
        actPlayer = {'number': actNumber, 'name': actName, 'points': actPoints, 'info': '', 'player_id': -1, 'global_pstat_id': -1}
        if res:
            found_ps = [ps for ps in global_ps_a if ps.player.id == player_id ]
            if len(global_ps_a) == 0:
                actPlayer['info'] = 'OK'
                actPlayer['player_id'] = player_id
            elif len(found_ps) == 0:
                actPlayer['info'] = 'Not found'
            else:
                actPlayer['info'] = 'OK'
                actPlayer['global_pstat_id'] = found_ps[0].id
                actPlayer['player_id'] = player_id
        else:
            actPlayer['info'] = 'No ID!'
        
        
        result['playerstats_a'].append(actPlayer)
    
    
    for iRow in range(0,tsettings.amount_players_report):
        actRow = iRow + tmp.team_b_start_row
        actNumber = ws[tmp.team_b_player_number_col+str(actRow)].value
        if actNumber is None:
            continue
        actName = ws[tmp.team_b_player_name_col+str(actRow)].value
        if actName is None or actName == '':
            continue
        actPoints = ws[tmp.team_b_player_points_col+str(actRow)].value
        if actPoints is None:
            actPoints = 0
        if actPoints > maxPoints:
            messages.append("#" + str(actNumber) + " score too high...")
            continue
        player_id, res = intTryParse(actName[actName.find("(")+1:actName.find(")")])
        actPlayer = {'number': actNumber, 'name': actName, 'points': actPoints, 'info': '', 'player_id': -1, 'global_pstat_id': -1}
        if res:
            found_ps = [ps for ps in global_ps_b if ps.player.id == player_id ]
            if len(global_ps_b) == 0:
                actPlayer['info'] = 'OK'
                actPlayer['player_id'] = player_id
            elif len(found_ps) == 0:
                actPlayer['info'] = 'Not found'
            else:
                actPlayer['info'] = 'OK'
                actPlayer['global_pstat_id'] = found_ps[0].id
                actPlayer['player_id'] = player_id
        else:
            actPlayer['info'] = 'No ID!'
        result['playerstats_b'].append(actPlayer)
    if len(local_ps_a) == 0 and len(local_ps_b) == 0:
        result['msg'] = 'Report upload successful'
    else:
        result['msg'] = 'Found existing stats. If you upload this report, you will overwrite them!'

    if len(messages) > 0:
        result['isError'] = True
        result['msg'] = 'Double check player scores. score is too high...'
        return result

    return result

def import_playerstats_game_report(game, upload_data):
    print('ENTER import_single_game_report')
    result = {'isError': False, 'msg': 'OK', 'playerstats_a': [], 'playerstats_b': [], 'score_team_a_halftime_1': 0, 'score_team_a_halftime_2': 0, 'score_team_a_penalty': 0, 'score_team_b_halftime_1': 0, 'score_team_b_halftime_2': 0, 'score_team_b_penalty': 0}

    players_a = [p for p in Player.objects.filter(team=game.team_a).all()]
    players_b = [p for p in Player.objects.filter(team=game.team_b).all()]
    global_ps_a = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_a, is_ranked=True).all()]
    global_ps_b = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_b, is_ranked=True).all()]

    local_ps_a = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, game=game, player__team=game.team_a, is_ranked=False).all()]
    local_ps_b = [ps for ps in PlayerStats.objects.filter(tournament_event=game.tournament_event, game=game, player__team=game.team_b, is_ranked=False).all()]

    #if local stats exists, we overwrite them: first subtract from global and then set new pstat
    for ps in local_ps_a:
        found_global_ps = [pst for pst in global_ps_a if pst.player.id == ps.player.id ]
        if len(found_global_ps) > 0:
            found_global_ps[0].score -= ps.score
            if found_global_ps[0].score < 0:
                found_global_ps[0].score = 0
            if found_global_ps[0].games_played > 0:
                found_global_ps[0].games_played -= 1
            found_global_ps[0].save()
        ps.delete()
    for ps in local_ps_b:
        found_global_ps = [pst for pst in global_ps_b if pst.player.id == ps.player.id ]
        if len(found_global_ps) > 0:          
            found_global_ps[0].score -= ps.score
            if found_global_ps[0].score < 0:
                found_global_ps[0].score = 0
            if found_global_ps[0].games_played > 0:
                found_global_ps[0].games_played -= 1
            found_global_ps[0].save()
        ps.delete()

    playerstats_bulk = []
    for ps in upload_data['playerstats_a']:
        if ps['player_id'] == -1:
            continue
        found_ps = [pst for pst in global_ps_a if pst.player.id == ps['player_id'] ]
        player = next(p for p in players_a if p.id == ps['player_id'] )
        if len(global_ps_a) == 0 or len(found_ps) == 0:           
            global_ps = PlayerStats(tournament_event=game.tournament_event,
             player_id=ps['player_id'],
             score=ps['points'],
             is_ranked=True,
             season_team_id=game.team_a.season_team_id,
             season_player_id=player.season_player_id,
             season_cup_tournament_id=game.tournament_event.season_cup_tournament_id,
             season_cup_german_championship_id=game.tournament_event.season_cup_german_championship_id,
             gbo_category_id=game.tournament_event.category.gbo_category_id)
            playerstats_bulk.append(global_ps)
        else:
            found_ps[0].score += ps['points']
            found_ps[0].games_played += 1
            found_ps[0].save()
        playerstats_bulk.append(PlayerStats(tournament_event=game.tournament_event,
         game=game,
         player_id=ps['player_id'],
         score=ps['points'],
         games_played=1,
         is_ranked=False,
         season_team_id=game.team_a.season_team_id,
         season_player_id=player.season_player_id,
         season_cup_tournament_id=game.tournament_event.season_cup_tournament_id,
         season_cup_german_championship_id=game.tournament_event.season_cup_german_championship_id,
         gbo_category_id=game.tournament_event.category.gbo_category_id)
        )
    
    
    for ps in upload_data['playerstats_b']:
        if ps['player_id'] == -1:
            continue
        found_ps = [pst for pst in global_ps_b if pst.player.id == ps['player_id'] ]
        player = next(p for p in players_b if p.id == ps['player_id'] )
        if len(global_ps_b) == 0 or len(found_ps) == 0:
            global_ps = PlayerStats(tournament_event=game.tournament_event,
             player_id=ps['player_id'],
             score=ps['points'],
             is_ranked=True,
             season_team_id=game.team_b.season_team_id,
             season_player_id=player.season_player_id,
             season_cup_tournament_id=game.tournament_event.season_cup_tournament_id,
             season_cup_german_championship_id=game.tournament_event.season_cup_german_championship_id,
             gbo_category_id=game.tournament_event.category.gbo_category_id)
            playerstats_bulk.append(global_ps)
        else:
            found_ps[0].score += ps['points']
            found_ps[0].games_played += 1
            found_ps[0].save()
        playerstats_bulk.append(PlayerStats(tournament_event=game.tournament_event,
         game=game,
         player_id=ps['player_id'],
         score=ps['points'],
         games_played=1,
         is_ranked=False,
         season_team_id=game.team_b.season_team_id,
         season_player_id=player.season_player_id,
         season_cup_tournament_id=game.tournament_event.season_cup_tournament_id,
         season_cup_german_championship_id=game.tournament_event.season_cup_german_championship_id,
         gbo_category_id=game.tournament_event.category.gbo_category_id)
        )

    if len(playerstats_bulk) > 0:
        PlayerStats.objects.bulk_create(playerstats_bulk)
    
    helper.recalc_global_pstats(game.tournament_event.id)
    return result


def intTryParse(value):
    try:
        return int(value), True
    except ValueError:
        return value, False


def import_game_report_excel():
    print('ENTER import_game_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)

    files = [f for f in os.listdir(os.path.join(settings.GAME_REPORT_DIR, 'post')) if f.endswith('.xlsx')]

    tourn = Tournament.objects.get(id=28)
    tsettings = TournamentSettings.objects.get(tournament=tourn)
    if not tsettings.game_report_template:
        return ''

    gr_template = tsettings.game_report_template

    for file in files:
        wb = load_workbook(filename = os.path.join(settings.GAME_REPORT_DIR, 'post', file))
        ws = wb.active

        game_number = int(ws['T4'].value[1:])
        #game = Game.objects.filter(tournament=tourn, id_counter=game_number)
        games = Game.objects.select_related("tournament", "tournament_event__category", "team_st_a__team", "team_st_b__team").prefetch_related(
            Prefetch("team_a__player_set", queryset=Player.objects.all(), to_attr="players"),
            Prefetch("team_b__player_set", queryset=Player.objects.all(), to_attr="players"),
            Prefetch("playerstats_set", queryset=PlayerStats.objects.select_related("tournament_event__category", "player__team", "teamstat").all(), to_attr="pstat")
        ).filter(tournament=tourn)
        all_games = [g for g in games.all()]
        game = next(g for g in all_games if g.id_counter==game_number)
        players_a = game.team_a.players #[pl for pl in game.team_a.player_set.all()]
        players_b = game.team_b.players #[pl for pl in game.team_b.player_set.all()]
        all_pstats = game.pstat
        pstats_a = [ps for ps in all_pstats if ps.player.team.id == game.team_st_a.team.id]
        pstats_b = [ps for ps in all_pstats if ps.player.team.id == game.team_st_b.team.id]
        for ps in pstats_a:
            ps.spin_try = 0
            ps.spin_success = 0
            ps.one_try = 0
            ps.one_success = 0
            ps.shooter_success = 0
            ps.shooter_try = 0
            ps.kempa_try = 0
            ps.kempa_success = 0
            ps.score = 0
        for ps in pstats_b:
            ps.spin_try = 0
            ps.spin_success = 0
            ps.one_try = 0
            ps.one_success = 0
            ps.shooter_success = 0
            ps.shooter_try = 0
            ps.kempa_try = 0
            ps.kempa_success = 0
            ps.score = 0

        global_ps_a = PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_a, is_ranked=True).all()
        global_ps_b = PlayerStats.objects.filter(tournament_event=game.tournament_event, player__team=game.team_b, is_ranked=True).all()


        score_a_1 = ws['M40'].value
        score_a_2 = ws['T40'].value
        score_a_p = ws['T44'].value
        if score_a_p is None:
            score_a_p = 0
        score_a_1_act = 0
        score_a_2_act = 0
        score_a_p_act = 0

        score_b_1 = ws['O40'].value
        score_b_2 = ws['V40'].value
        score_b_p = ws['V44'].value
        if score_b_p is None:
            score_b_p = 0
        score_b_1_act = 0
        score_b_2_act = 0
        score_b_p_act = 0
        error_found = False
        # set 1 team a
        set1_start_row = 13
        set1_end_row = 32

        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            print(str(ws['J' + str(act_row)].value))
            print(str(ws['B13'].value))
            if not ws['J' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['J' + str(act_row)].value)
                player = next(pl for pl in players_a if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_a if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_a if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_a_1_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            if not ws['L' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['L' + str(act_row)].value)
                player = next(pl for pl in players_b if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_b if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_b if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_b_1_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0

        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            print(str(ws['M' + str(act_row)].value))
            print(str(ws['B13'].value))
            if not ws['M' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['M' + str(act_row)].value)
                player = next(pl for pl in players_a if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_a if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_a if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_a_1_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            if not ws['O' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['O' + str(act_row)].value)
                player = next(pl for pl in players_b if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_b if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_b if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_b_1_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0

        if score_a_1_act != score_a_1:
            print('ERROR Halftime 1')
            error_found = True
        if score_b_1_act != score_b_1:
            print('ERROR Halftime 1')
            error_found = True
        
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            print(str(ws['Q' + str(act_row)].value))
            print(str(ws['B13'].value))
            if not ws['Q' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['Q' + str(act_row)].value)
                print(number)
                player = next((pl for pl in players_a if pl.number == number), None)
                if player:
                    pstat = next((ps for ps in pstats_a if ps.player_id == player.id), None)
                    pstat_gl = next((ps for ps in global_ps_a if ps.player_id == player.id), None)
                    if pstat and pstat_gl:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_a_2_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                        error_found = True
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))          
            if act_points == 2:
                act_points = 0
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            if not ws['S' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['S' + str(act_row)].value)
                player = next(pl for pl in players_b if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_b if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_b if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_b_2_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
            if act_points == 2:
                act_points = 0
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            print(str(ws['M' + str(act_row)].value))
            print(str(ws['B13'].value))
            if not ws['T' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['T' + str(act_row)].value)
                player = next(pl for pl in players_a if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_a if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_a if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_a_2_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0
        act_points = 0
        for iRow in range(1,21):
            act_points += 1
            act_row = set1_start_row + iRow
            if not ws['V' + str(act_row)].value is None:
                #player of a scored
                number = int(ws['V' + str(act_row)].value)
                player = next(pl for pl in players_b if pl.number == number)
                if player:
                    pstat = next(ps for ps in pstats_b if ps.player_id == player.id)
                    pstat_gl = next(ps for ps in global_ps_b if ps.player_id == player.id)
                    if pstat:
                        if act_points == 1:
                            pstat.one_success += 1
                            pstat.one_try += 1
                            pstat_gl.one_success += 1
                            pstat_gl.one_try += 1
                        else:
                            pstat.spin_success += 1
                            pstat.spin_try += 1
                            pstat_gl.spin_success += 1
                            pstat_gl.spin_try += 1
                        pstat.score += act_points
                        pstat_gl.score += act_points
                        score_b_2_act += act_points
                        if act_points == 1:
                            act_points = 0
                    else:
                        print('No PlayerStat found')
                else:
                    error_found = True
                    print('Playernumber not found '+ str(number))
            if act_points == 2:
                act_points = 0
        
        if score_a_2_act != score_a_2:
            print('ERROR Halftime 2')
            error_found = True
        if score_b_2_act != score_b_2:
            print('ERROR Halftime 2')
            error_found = True
              
        # penalty
        act_points = 0
        for iRow in range(11,16):
            print(str(ws.cell(column=iRow, row=48).value))
            if not ws.cell(column=iRow, row=48).value is None and not ws.cell(column=iRow, row=50).value is None:
                #player of a scored
                number = int(ws.cell(column=iRow, row=48).value)
                print(number)
                score = int(ws.cell(column=iRow, row=50).value)
                if not score is None:
                    player = next(pl for pl in players_a if pl.number == number)
                    if player:
                        pstat = next((ps for ps in pstats_a if ps.player_id == player.id), None)
                        pstat_gl = next((ps for ps in global_ps_a if ps.player_id == player.id), None)
                        if pstat and pstat_gl:
                            if score == 1:
                                pstat.one_success += 1
                                pstat.one_try += 1
                                pstat_gl.one_success += 1
                                pstat_gl.one_try += 1
                            else:
                                pstat.spin_success += 1
                                pstat.spin_try += 1
                                pstat_gl.spin_success += 1
                                pstat_gl.spin_try += 1
                            pstat.score += score
                            pstat_gl.score += score
                            score_a_p_act += score
                        else:
                            print('No PlayerStat found')
                            error_found = True
                    else:
                        error_found = True
                        print('Playernumber not found '+ str(number))
            if not ws.cell(column=iRow, row=52).value is None and not ws.cell(column=iRow, row=54).value is None:
                #player of a scored
                number = int(ws.cell(column=iRow, row=52).value)
                score = int(ws.cell(column=iRow, row=54).value)
                if not score is None:
                    player = next(pl for pl in players_b if pl.number == number)
                    if player:
                        pstat = next((ps for ps in pstats_b if ps.player_id == player.id), None)
                        pstat_gl = next((ps for ps in global_ps_b if ps.player_id == player.id), None)
                        if pstat and pstat_gl:
                            if act_points == 1:
                                pstat.one_success += 1
                                pstat.one_try += 1
                                pstat_gl.one_success += 1
                                pstat_gl.one_try += 1
                            else:
                                pstat.spin_success += 1
                                pstat.spin_try += 1
                                pstat_gl.spin_success += 1
                                pstat_gl.spin_try += 1
                            pstat.score += score
                            pstat_gl.score += score
                            score_b_p_act += score
                        else:
                            print('No PlayerStat found')
                            error_found = True
                    else:
                        error_found = True
                        print('Playernumber not found '+ str(number))
        if score_a_p_act != score_a_p:
            print('ERROR Penalty')
            error_found = True
        if score_b_p_act != score_b_p:
            print('ERROR Penalty')
            error_found = True

        for psg in global_ps_a:
            psg.games_played += 1
        for psg in global_ps_b:
            psg.games_played += 1
        if not error_found:
            # save
            #['score','kempa_try', 'kempa_success','kempa_try', 'spin_success','spin_try', 'shooter_success','shooter_try', 'one_try', 'one_success']
            PlayerStats.objects.bulk_update(pstats_a, fields=['score','spin_success','spin_try', 'one_try', 'one_success'])
            PlayerStats.objects.bulk_update(pstats_b, fields=['score','spin_success','spin_try', 'one_try', 'one_success', 'games_played'])
            PlayerStats.objects.bulk_update(global_ps_a, fields=['score','spin_success','spin_try', 'one_try', 'one_success'])
            PlayerStats.objects.bulk_update(global_ps_b, fields=['score','spin_success','spin_try', 'one_try', 'one_success'])

        tstage = game.tournament_state.tournament_stage
        
        ws["T4"] = game.tournament_event.category.abbreviation + str(game.id_counter)
        