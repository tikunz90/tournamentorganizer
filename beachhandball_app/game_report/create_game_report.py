from beachhandball_app.models.Tournaments import TournamentSettings
from beachhandball_app.models.choices import CATEGORY_CHOICES
from beachhandball_app.models.Game import Game

import os
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from django.conf import settings

def create_pregame_report_excel(game):
    print('ENTER create_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)

    filename_template = 'game_report_template.xlsx'
    filename = 'PreGame_' + game.tournament_event.category.abbreviation + str(game.id) + '.xlsx'

    fullfilepath_template = os.path.join(settings.GAME_REPORT_DIR, filename_template)
    fullfilepath_report = os.path.join(settings.GAME_REPORT_DIR, filename)
    # copy template
    copyfile(fullfilepath_template, fullfilepath_report)

    if os.path.isfile(fullfilepath_report):
        print('file exists')
        wb = load_workbook(filename = fullfilepath_report)
        ws = wb.active
        #img = Image(os.path.join(settings.GAME_REPORT_DIR,'dhb_logo.png'))
        #ws_copy = wb.copy_worksheet(ws)
        #ws.add_image(img, 'T1')
        # game id
        tsettings = TournamentSettings.objects.get(tournament=game.tournament_event.tournament)
        ws["T4"] = game.tournament_event.category.abbreviation + str(game.id)
        
        # category
        if game.tournament_event.category.category == CATEGORY_CHOICES[0][0]:
            ws["M3"] = 'X'
        elif game.tournament_event.category.category == CATEGORY_CHOICES[1][0]:
            ws["P3"] = 'X'

        # court
        ws["R8"] = game.court.number

        # team names
        ws["B8"] = game.team_st_a.team.name
        ws["G8"] = game.team_st_b.team.name

        # Date and Time
        date = game.starttime.strftime("%d.%m.%Y")
        time = game.starttime.strftime("%H:%M")
        ws["C10"] = date
        ws["F10"] = time
        max_num_player = tsettings.amount_players_report
        row_start_a = 13
        act_player_counter = 1
        for player in game.team_st_a.team.player_set.all():
            if act_player_counter > max_num_player:
                break
            if not player.is_active:
                continue
            ws["A"+str(row_start_a)]=player.number
            ws["B"+str(row_start_a)]=player.name + ", " + player.first_name
            act_player_counter = act_player_counter + 1
            row_start_a = row_start_a + 1
        row_start_b = 30
        act_player_counter = 1
        for player in game.team_st_b.team.player_set.all():
            if act_player_counter > max_num_player:
                break
            if not player.is_active:
                continue
            ws["A"+str(row_start_b)]=player.number
            ws["B"+str(row_start_b)]=player.name + ", " + player.first_name
            act_player_counter = act_player_counter + 1
            row_start_b = row_start_b + 1
        max_num_coaches = 2
        row_start_coach_a = 25
        for coach in game.team_st_a.team.coach_set.all()[:max_num_coaches]:
            ws["B"+str(row_start_coach_a)]=coach.name + ", " + coach.first_name
            row_start_coach_a = row_start_coach_a + 1
        row_start_coach_b = 42
        for coach in game.team_st_b.team.coach_set.all()[:max_num_coaches]:
            ws["B"+str(row_start_coach_b)]=coach.name + ", " + coach.first_name
            row_start_coach_b = row_start_coach_b + 1

        wb.save(fullfilepath_report)

        return fullfilepath_report, filename
    return ''


def create_all_tstate_pregame_report_excel(tstate):
    print('ENTER create_report_excel')
    print('game_report DIR: ' + settings.GAME_REPORT_DIR)

    filename_template = 'game_report_template.xlsx'
    filename = 'PreGame_ALL_'  + str(tstate.name) + '.xlsx'

    fullfilepath_template = os.path.join(settings.GAME_REPORT_DIR, filename_template)
    fullfilepath_report = os.path.join(settings.GAME_REPORT_DIR, filename)
    # copy template
    copyfile(fullfilepath_template, fullfilepath_report)

    if os.path.isfile(fullfilepath_report):
        print('file exists')
        tsettings = TournamentSettings.objects.get(tournament=tstate.tournament_event.tournament)
        games = Game.objects.filter(tournament_state=tstate)
        wb = load_workbook(filename = fullfilepath_report)
        ws_origin = wb.active
        #img = Image(os.path.join(settings.GAME_REPORT_DIR,'dhb_logo.png'))
        for game in games:
            
            ws_game = wb.copy_worksheet(ws_origin)
            ws_game.title = game.tournament_event.category.abbreviation + str(game.id)
            # game id
            ws_game["T4"] = game.tournament_event.category.abbreviation + str(game.id)
            
            # category
            if game.tournament_event.category.category == CATEGORY_CHOICES[0][0]:
                ws_game["M3"] = 'X'
            elif game.tournament_event.category.category == CATEGORY_CHOICES[1][0]:
                ws_game["P3"] = 'X'

            # court
            ws_game["R8"] = game.court.number

            # team names
            ws_game["B8"] = game.team_st_a.team.name
            ws_game["G8"] = game.team_st_b.team.name

            # Date and Time
            date = game.starttime.strftime("%d.%m.%Y")
            time = game.starttime.strftime("%H:%M")
            ws_game["C10"] = date
            ws_game["F10"] = time
            max_num_player = tsettings.amount_players_report
            row_start_a = 13
            act_player_counter = 1
            for player in game.team_st_a.team.player_set.all():
                if act_player_counter > max_num_player:
                    break
                if not player.is_active:
                    continue
                ws_game["A"+str(row_start_a)]=player.number
                ws_game["B"+str(row_start_a)]=player.name + ", " + player.first_name
                act_player_counter = act_player_counter + 1
                row_start_a = row_start_a + 1
            row_start_b = 30
            act_player_counter = 1
            for player in game.team_st_b.team.player_set.all():
                if act_player_counter > max_num_player:
                    break
                if not player.is_active:
                    continue
                ws_game["A"+str(row_start_b)]=player.number
                ws_game["B"+str(row_start_b)]=player.name + ", " + player.first_name
                act_player_counter = act_player_counter + 1
                row_start_b = row_start_b + 1
            max_num_coaches = 2
            row_start_coach_a = 25
            for coach in game.team_st_a.team.coach_set.all()[:max_num_coaches]:
                ws_game["B"+str(row_start_coach_a)]=coach.name + ", " + coach.first_name
                row_start_coach_a = row_start_coach_a + 1
            row_start_coach_b = 42
            for coach in game.team_st_b.team.coach_set.all()[:max_num_coaches]:
                ws_game["B"+str(row_start_coach_b)]=coach.name + ", " + coach.first_name
                row_start_coach_b = row_start_coach_b + 1
        wb.remove_sheet(ws_origin)
        wb.save(fullfilepath_report)

        return fullfilepath_report, filename
    return ''